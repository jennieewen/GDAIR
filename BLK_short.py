import os
import openpyxl
from openpyxl import load_workbook

def Report_BLK_short(reportDate, analyzeDate, batch):

	appPath = os.path.dirname(os.path.abspath(__file__))
	templateFile= appPath + '/Template/Template-BTX-BLK-Agilent.xlsx'
	inputFile = appPath + '/BLK/Source/epatemp.txt'
	outputFile = appPath + "/BLK/Target/BTX-BLK-" + reportDate.replace("/","") + "-Agilent.xlsx"
	sheetName = "Report"

	sheetStartRow = 18


	class TemplateData(object):
		def __init__(self, CONSTITUENT='', result=''):
			self.CONSTITUENT = CONSTITUENT # avoid using Python keywords where possible
			self.Result = Result

	wb = load_workbook(templateFile)
	sheetname = sheetName
	ws = wb[sheetname]

	TemplateData_list = []
	i = sheetStartRow

	for row in ws.rows:
		if ws["A"+ str(i)].value != None:
			CONSTITUENT = ws["A"+ str(i)].value
			Result = ws["H" + str(i)].value
			r = TemplateData(CONSTITUENT, Result)
			TemplateData_list .append(r)
		
		else:
			break
		i += 1

	class Item():
		def __init__(self, CONSTITUENT='', Result=''): #Mark = ug/cu M 
			self.CONSTITUENT = CONSTITUENT 
			self.Result = Result
			
	lines = []
	startpos = 0
	endpos = 0
	i = 1
	with open(inputFile, 'r') as fh:
		for line in fh:
			lines.append(line)
			if 'Target Compounds' in line:
				startpos = i
			if 	'qualifier out of range' in line:
				endpos = i-3
			i += 1 	
		
	date_lines = lines[startpos:endpos]

	items = []
	for date_line in date_lines:
		constituent = date_line[7:33].strip()
		result = date_line[56:64].strip()
		
		r = Item(constituent, result,)
		items.append(r)

	for templatedata in TemplateData_list:
		name = templatedata.CONSTITUENT
		for item in items:
			if item.CONSTITUENT.strip() == name.strip():
				templatedata.Result = item.Result
				break
			

	in_file = open(templateFile, 'rb')
	indata = in_file.read()

	out_file = open(outputFile, 'wb+')
	out_file.write(indata)

	out_file.close()
	in_file.close()

	wb = openpyxl.load_workbook(outputFile)
	sheetname = sheetName
	ws = wb.active

	ws["J4"].value = reportDate
	ws["J5"].value = analyzeDate
	ws["J7"].value = batch

	k = sheetStartRow
	i = 0
	for stuff in TemplateData_list:
		ws["H"+ str(k+i)].value = stuff.Result
		# use formular: =IF(MID(H17,1,1)="N",H17,H17/24.45*E17)
		ws["I"+ str(k+i)].value = '=IF(MID(H{0},1,1)="N",H{0},H{0}/24.45*E{0})'.format(k+i)   	#stuff.Mark
		i += 1

	wb.save(outputFile)
	wb.close()
	print("Sucessfully generated file " + outputFile)
