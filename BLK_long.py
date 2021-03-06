import os
import openpyxl
from openpyxl import load_workbook

def Report_BLK_long(reportDate, analyzeDate, batch):

	appPath = os.path.dirname(os.path.abspath(__file__))
	static_file = appPath + '/Template/StaticData.xlsx'
	templateFile= appPath + '/Template/Template-TO15-BLK-Agilent.xlsx'
	inputFile = appPath + '/BLK/Source/epatemp.txt'
	outputFile = appPath + "/BLK/Target/TO15-blk-" + reportDate.replace("/","") + "-Agilent.xlsx"
	sheetName = "Report"

	sheetStartRow = 17
	sheetDataRows = 33
	sheetPageRows = 54

	class StaticData(object):
		def __init__(self, CONSTITUENT='',MW='', CAS='', PQL = '', Note = '', ORDERBY = ''):
			self.CONSTITUENT = CONSTITUENT # avoid using Python keywords where possible
			self.MW = MW
			self.CAS = CAS
			self.PQL = PQL
			self.Note = Note
			self.ORDERBY = ORDERBY

	wb = load_workbook(static_file)
	sheetname = sheetName
	ws = wb[sheetname]

	staticdata_list = []
	i = 2
	for row in ws.rows:
		CONSTITUENT = ws["A"+ str(i)].value
		MW = ws["B" + str(i)].value 
		CAS = ws["C" + str(i)].value
		PQL = ws["D" + str(i)].value
		Note = ws["E" + str(i)].value 
		ORDERBY = ws["F" + str(i)].value
		r = StaticData(CONSTITUENT, MW, CAS, PQL, Note, ORDERBY)
		staticdata_list.append(r)
		i += 1
	
	class Item():
		def __init__(self, CONSTITUENT='', MW='', CAS='', PQL='', Result='', Mark='', Note='', ORDERBY=''): #Mark = ug/cu M 
			self.CONSTITUENT = CONSTITUENT 
			self.MW = MW
			self.CAS = CAS
			self.PQL = PQL
			self.Result = Result
			self.Mark = Mark
			self.Note = Note
			self.ORDERBY = ORDERBY

	lines = []
	startpos = 0
	endpos = 0
	i = 1
	with open(inputFile, 'r') as fh:
		for line in fh:
			lines.append(line)
			if 'Target Compounds' in line:
				startpos = i
			if 	'qualifier out of range' in line:
				endpos = i-3
			i += 1 	

				
	data_lines = lines[startpos:endpos]

	items = []
	for date_line in data_lines:
		constituent = date_line[7:33].strip()
		result = date_line[56:64].strip()
		
		r = Item(constituent, "", "", "", result, "", "")
		items.append(r)

	for item in items:
		if item.CONSTITUENT != None: 
			for data in staticdata_list:	
				if data.CONSTITUENT != None and item.CONSTITUENT.strip() == data.CONSTITUENT.strip():
					item.MW = data.MW
					item.CAS = data.CAS
					item.PQL = data.PQL
					item.Note = data.Note
					item.ORDERBY = data.ORDERBY

	sorted_list = sorted(items, key = lambda p: p.ORDERBY)		
		
	in_file = open(templateFile, 'rb')
	indata = in_file.read()

	out_file = open(outputFile, 'wb+')
	out_file.write(indata)

	out_file.close()
	in_file.close()


	wb = openpyxl.load_workbook(outputFile)
	sheetname = sheetName
	ws = wb.active

	ws["J4"].value = reportDate
	ws["J5"].value = analyzeDate
	ws["J7"].value = batch

	k = sheetStartRow
	i = 0
	for stuff in sorted_list:
		if i == sheetDataRows:
			k += sheetPageRows 
			i = 0
		ws["A"+ str(k+i)].value = stuff.CONSTITUENT
		ws["E"+ str(k+i)].value = stuff.MW
		ws["F"+ str(k+i)].value = stuff.CAS
		ws["G"+ str(k+i)].value = stuff.PQL
		ws["H"+ str(k+i)].value = stuff.Result
		# use formular: =IF(MID(H17,1,1)="N",H17,H17/24.45*E17)
		ws["I"+ str(k+i)].value = '=IF(MID(H{0},1,1)="N",H{0},H{0}/24.45*E{0})'.format(k+i)   	#stuff.Mark
		ws["J"+ str(k+i)].value = stuff.Note
		i += 1

	wb.save(outputFile)
	wb.close()
	print("Sucessfully generated file " + outputFile)
