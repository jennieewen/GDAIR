import os
import openpyxl
from openpyxl import load_workbook

def Report_BS_short(reportDate, analyzeDate, batch):

	appPath = os.path.dirname(os.path.abspath(__file__))
	templateFile= appPath + '/Template/Template-BTX-BS-Agilent.xlsx'
	inputBSFile = appPath + '/BS/Source/epatemp-bs.txt'
	inputBSDFile = appPath + '/BS/Source/epatemp-bsd.txt'
	outputFile = appPath + "/BS/Target/BTX-BS-" + reportDate.replace("/","") + "-Agilent.xlsx"
	sheetName = "Report"
	sheetStartRow = 17

	class BS(object):
		def __init__(self, constituent, bs):
			self.constituent = constituent
			self.bs = bs
			
	class BSD(object):
		def __init__(self, constituent, bsd):
			self.constituent = constituent
			self.bsd = bsd

	bs_lines = []
	startpos = 0
	endpos = 0
	i = 1
	with open(inputBSFile, 'r') as fh:
		for line in fh:
			bs_lines.append(line)
			if 'Target Compounds' in line:
				startpos = i
			if 	'qualifier out of range' in line:
				endpos = i-3
			i += 1 	

	data_bs_lines = bs_lines[startpos:endpos]

	bs_items = []
	for data_bs_line in data_bs_lines:
		constituent = data_bs_line[7:33].strip()
		bs = data_bs_line[59:64].strip()

		r = BS(constituent, bs)
		bs_items.append(r)
			

	bsd_lines = []
	startpos = 0
	endpos = 0
	i = 1
	with open(inputBSDFile, 'r') as fd:
		for line in fd:
			bsd_lines.append(line)
			if 'Target Compounds' in line:
				startpos = i
			if 	'qualifier out of range' in line:
				endpos = i-3
			i += 1 	

	data_bsd_lines = bsd_lines[startpos:endpos]

	bsd_items = []
	for data_bsd_line in data_bsd_lines:
		constituent = data_bsd_line[7:33].strip()
		bsd = data_bsd_line[59:64].strip()

		# put items list into class 
		d = BSD(constituent,bsd)
		bsd_items.append(d)
		
	in_file = open(templateFile, 'rb')
	indata = in_file.read()

	out_file = open(outputFile, 'wb+')
	out_file.write(indata)

	out_file.close()
	in_file.close()


	wb = openpyxl.load_workbook(outputFile)
	ws = wb.active

	# fill static fields
	ws["J2"].value = reportDate
	ws["J3"].value = analyzeDate
	ws["J5"].value = batch

	i = sheetStartRow
	for row in ws.rows:
		name = ws["A"+ str(i)].value
		for item in bs_items:
			if item.constituent.strip() == name:
				ws["F"+ str(i)].value = item.bs
				break
				
		for item in bsd_items:
			if item.constituent.strip() == name:
				ws["H"+ str(i)].value = item.bsd
				break
		i += 1
		

	wb.save(outputFile)
	wb.close()
	print("Sucessfully generated file " + outputFile)
